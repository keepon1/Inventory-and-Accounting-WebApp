from react_work import models
from reportlab.platypus import SimpleDocTemplate, Table as RLTable, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table as XLTable, TableStyleInfo
from io import BytesIO, StringIO
import base64
import csv
from datetime import datetime


class PDF():
    def __init__(self, data, location, start, end, user):
        self.data = data
        self.location = location
        self.start = start
        self.end = end
        self.user = user

    def generate_item_pdf(self):
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, leftMargin=40, rightMargin=40, topMargin=40, bottomMargin=40)
        styles = getSampleStyleSheet()
        flowables = []

        title = Paragraph(f"Inventory Items Report for {self.location}", styles['Heading2'])
        flowables.append(title)
        flowables.append(Spacer(1, 12))

        headers = ['Item Code', 'Item Name', 'Category', 'Brand', 'Quantity', 'Cost', 'Price', 'Total Value']
        table_data = [headers]

        for item in self.data:
            qty = item.get('quantity', 0) or 0
            cost = item.get('purchase_price', 0.0) or 0.0
            price = item.get('sales_price', 0.0) or 0.0
            total_value = qty * cost
            row = [
                item.get('code', ''),
                item.get('item_name', ''),
                item.get('category__name', ''),
                item.get('brand__name', ''),
                str(qty),
                f"{cost:.2f}",
                f"{price:.2f}",
                f"{total_value:.2f}",
            ]
            table_data.append(row)

        col_widths = [50, 120, 90, 90, 50, 60, 60, 70]

        rl_table = RLTable(table_data, colWidths=col_widths, repeatRows=1)
        rl_style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#CCCCCC')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('ALIGN', (1, 1), (1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ])
        rl_table.setStyle(rl_style)

        flowables.append(rl_table)
        doc.build(flowables)

        buffer.seek(0)

        today = datetime.now()

        return {
            'file': base64.b64encode(buffer.getvalue()).decode('utf-8'),
            'filename': f'inventory_items_{self.location}_{today}.pdf',
            'type': 'application/pdf'
        }


class XLSX():
    def __init__(self, data, location, start, end, user):
        self.data = data
        self.location = location
        self.start = start
        self.end = end
        self.user = user

    def generate_item_xlsx(self):
        wb = Workbook()

        active_sheet = wb.active
        if active_sheet is not None:
            wb.remove(active_sheet)
        ws = wb.create_sheet(title="Inventory Items")

        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

        ws['A1'] = f"Inventory Items Report for {self.location}"
        ws.merge_cells('A1:J1')
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center')

        headers = ['Item Code', 'Item Name', 'Category', 'Brand', 'Model', 'Quantity', 'Unit', 'Cost', 'Price', 'Total Value']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        for row_index, item in enumerate(self.data, start=4):
            ws.cell(row=row_index, column=1, value=item.get('code', ''))
            ws.cell(row=row_index, column=2, value=item.get('item_name', ''))
            ws.cell(row=row_index, column=3, value=item.get('category__name', ''))
            ws.cell(row=row_index, column=4, value=item.get('brand__name', ''))
            ws.cell(row=row_index, column=5, value=item.get('model', ''))
            ws.cell(row=row_index, column=6, value=item.get('quantity', 0))
            ws.cell(row=row_index, column=7, value=item.get('unit__suffix', ''))
            ws.cell(row=row_index, column=8, value=item.get('purchase_price', 0.0))
            ws.cell(row=row_index, column=9, value=item.get('sales_price', 0.0))
            total_value = (item.get('quantity', 0) or 0) * (item.get('cost', 0.0) or 0.0)
            ws.cell(row=row_index, column=10, value=total_value)

        last_row = max(3, 3 + len(self.data))
        ref = f"A3:J{last_row}"
        table = XLTable(displayName="ItemsTable", ref=ref)
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                               showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        table.tableStyleInfo = style
        ws.add_table(table)

        ws.sheet_view.showGridLines = False

        for idx in range(1, ws.max_column + 1):
            column_letter = get_column_letter(idx)
            max_length = 0
            for cell in ws[column_letter]:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column_letter].width = max_length + 2

        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        today = datetime.now()

        return {
            'file': base64.b64encode(excel_file.getvalue()).decode('utf-8'),
            'filename': f'inventory_items_{self.location}_{today}.xlsx',
            'type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        }


class CSV():
    def __init__(self, data, location, start, end, user):
        self.data = data
        self.location = location
        self.start = start
        self.end = end
        self.user = user

    def generate_item_csv(self):
        buffer = StringIO()

        writer = csv.writer(buffer)
        headers = ['Item Code', 'Item Name', 'Category', 'Brand', 'Model', 'Quantity', 'Unit', 'Cost', 'Price', 'Total Value']
        writer.writerow(headers)
        for item in self.data:
            qty = item.get('quantity', 0) or 0
            cost = item.get('purchase_price', 0.0) or 0.0
            total_value = qty * cost
            writer.writerow([
                item.get('code', ''),
                item.get('item_name', ''),
                item.get('category__name', ''),
                item.get('brand__name', ''),
                item.get('model', ''),
                qty,
                item.get('unit__suffix', ''),
                f"{cost:.2f}",
                f"{item.get('sales_price', 0.0):.2f}",
                f"{total_value:.2f}",
            ])

        csv_bytes = buffer.getvalue().encode('utf-8')
        buffer.close()

        today = datetime.now()

        return {
            'file': base64.b64encode(csv_bytes).decode('utf-8'),
            'filename': f'inventory_items_{self.location}_{today}.csv',
            'type': 'text/csv'
        }